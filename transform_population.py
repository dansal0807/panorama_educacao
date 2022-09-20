import requests
import os
from zipfile import ZipFile
import pandas as pd
import ast
import openpyxl
import unidecode

dir = 'população_estimada_IBGE'
cwd = os.getcwd()
new_dir = os.path.join(cwd,dir)

if dir not in os.listdir():
    os.mkdir(new_dir)

sheetnames = ['1-Faixa etária (entre a população de zero a 18 anos de idade) - Zero a três anos de idade',
              '2-Faixa etária (entre a população de zero a 18 anos de idade) - Quatro a cinco anos de idade',
              '3-Faixa etária (entre a população de zero a 18 anos de idade) - Seis a 14 anos de idade',
              '4-Faixa etária (entre a população de zero a 18 anos de idade) - 15 a 17 anos de idade',
              '5-Faixa etária (entre a população de zero a 18 anos de idade) - 18 anos de idade',
              '6-População total',
              '7-População entre zero e 19 anos de idade',
              '8-População entre zero e 18 anos de idade']
   

def get_xlsx_columns(wb):
    #wb: arquivo excel em openpyxl.
    for sheet in wb.worksheets:
        columns = []
        for i in range(1,sheet.max_column+1):
            cell = sheet.cell(row=3, column=i).value
            columns.append(cell)

    return columns

def opening_file_to_df(aimed_dir='arquivos_extraidos'):
    excel = [ os.path.join(root,file)  for root, dirs, files in os.walk(aimed_dir, topdown=False) for file in files if 'População estimada pelo IBGE segundo faixas etárias.xlsx' in file ][0]
    dfs = []

    wb = openpyxl.load_workbook(excel) 
    res = len(wb.sheetnames)

    columns = get_xlsx_columns(wb)
    for i in range(1,res):
        df = pd.read_excel(excel, sheet_name=i)
        df.columns = columns
        dfs.append(df)
    
    return dfs
    
#transform .xls to .xlsx
def getting_cities_code(aimed_dir = "arquivos_extraidos"):
    ibge_excel = [ os.path.join(root,file)  for root, dirs, files in os.walk(aimed_dir, topdown=False) for file in files if 'RELATORIO_DTB_BRASIL_MUNICIPIO-ibge.xlsx' in file ][0]
    ibge_df = pd.read_excel(ibge_excel)
    ibge_rj = ibge_df[ibge_df["Nome_UF"] == "Rio de Janeiro"]
    ibge_codes_rj = ibge_rj['Código Município Completo'].tolist()
    #Adicionando o Estado do Rio de Janeiro como um todo
    ibge_codes_rj.append(33)
    ibge_codes_rj.sort()

    return ibge_codes_rj
    
def filtering_csvs_by_code(dfs):
    ibge_codes_rj = getting_cities_code()
    #lista de dataframes das sheets do arquivos xlsx.
    c = 0
    for df in dfs:
        df = df[df['Código IBGE'].isin(ibge_codes_rj)]
        
        if 'id' not in df.columns:
            df.reset_index(drop=True,inplace=True)
            df['id'] = df.index + 1
            columns = list(df)
            columns.insert(0, columns.pop(columns.index('id')))
            df = df.loc[:,columns]
            df.set_index('id',inplace=True)
        
        columns = [ column.lower().split() for column in df.columns ]
        columns = [ unidecode.unidecode("_".join(column)) for column in columns ]

        for i in range(len(columns)):
            if str.isdigit(columns[i]):
                columns[i] = "ano_" + columns[i]

        df.columns = columns

        df.to_csv(f'{os.path.join(new_dir,sheetnames[c])}.csv', sep=';', encoding='utf-8-sig')   
        c +=1     
    

def main():
    dfs = opening_file_to_df()
    filtering_csvs_by_code(dfs)

main()
